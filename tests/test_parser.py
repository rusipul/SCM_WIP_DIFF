import openpyxl
import pytest

from scm_wip_diff.parser import find_report_anchor, get_stage_groups, ReportFormatError
from scm_wip_diff.parser import build_column_labels
from scm_wip_diff.parser import parse_wip_sheet
from scm_wip_diff.parser import REPORT_TITLE

FIXTURE_260721 = "tests/fixtures/260721 GTK WIP.xlsx"
FIXTURE_260722 = "tests/fixtures/260722 GTK WIP.xlsx"


def test_find_report_anchor_locates_title_row():
    wb = openpyxl.load_workbook(FIXTURE_260721, data_only=True)
    ws = wb[wb.sheetnames[0]]

    anchor = find_report_anchor(ws)

    assert anchor == 2


def test_find_report_anchor_raises_when_title_missing():
    wb = openpyxl.Workbook()
    ws = wb.active

    with pytest.raises(ReportFormatError):
        find_report_anchor(ws)


def test_get_stage_groups_matches_merged_header_cells():
    wb = openpyxl.load_workbook(FIXTURE_260721, data_only=True)
    ws = wb[wb.sheetnames[0]]
    anchor = find_report_anchor(ws)

    groups = get_stage_groups(ws, anchor)

    assert groups["전공정"] == [9, 10, 11]
    assert groups["후공정"] == list(range(12, 29))
    assert groups["완료"] == [29, 30]


def test_build_column_labels_combines_two_header_rows():
    wb = openpyxl.load_workbook(FIXTURE_260721, data_only=True)
    ws = wb[wb.sheetnames[0]]
    anchor = find_report_anchor(ws)

    labels = build_column_labels(ws, anchor)

    assert labels[9] == "Saw"
    assert labels[10] == "Die Mount"
    assert labels[11] == "Wire Bond"
    assert labels[21] == "Ball Mount"
    assert labels[29] == "TR Stock"
    assert labels[30] == "Non TR Stock"


def test_parse_wip_sheet_reads_all_lot_rows_and_stops_before_next_report():
    parsed = parse_wip_sheet(FIXTURE_260721)

    assert len(parsed.lots) == 174


def test_parse_wip_sheet_extracts_process_values_for_first_data_row():
    parsed = parse_wip_sheet(FIXTURE_260721)
    key = ("SNF08N5623340E", "1B4686", "TMP1200D", 0)

    assert parsed.rows[key] == 7
    assert parsed.lots[key][9] == 0     # Saw
    assert parsed.lots[key][29] == 0    # TR Stock
    assert parsed.lots[key][30] == 24627  # Non TR Stock


def test_parse_wip_sheet_raises_when_stage_group_missing(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=1, value=REPORT_TITLE)

    # Group-header row (anchor + 1 = 3). Only "전공정" and "후공정" are
    # given merged cells; "완료" is intentionally omitted to simulate
    # format drift in the source file.
    ws.merge_cells(start_row=3, start_column=9, end_row=3, end_column=11)
    ws.cell(row=3, column=9, value="전공정")
    ws.merge_cells(start_row=3, start_column=12, end_row=3, end_column=28)
    ws.cell(row=3, column=12, value="후공정")

    path = tmp_path / "broken.xlsx"
    wb.save(path)

    with pytest.raises(ReportFormatError):
        parse_wip_sheet(str(path))


def test_parse_wip_sheet_disambiguates_duplicate_keys_by_occurrence():
    parsed = parse_wip_sheet(FIXTURE_260721)

    first_key = ("", "1B468601", "TMP1201D", 0)
    second_key = ("", "1B468601", "TMP1201D", 1)

    assert parsed.rows[first_key] == 18
    assert parsed.rows[second_key] == 19


def test_parse_wip_sheet_populates_sheet_name_and_number_format():
    parsed = parse_wip_sheet(FIXTURE_260721)

    assert parsed.sheet_name == "gtk3387"
    assert parsed.value_number_format == "#,##0"

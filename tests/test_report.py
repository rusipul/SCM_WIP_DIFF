import hashlib
import os
import shutil

import openpyxl
import pytest

from scm_wip_diff.atx_parser import parse_atx_wip_sheet
from scm_wip_diff.comparator import compare_lots, compare_stage_summary
from scm_wip_diff.report import build_highlighted_today_file, build_variance_report, derive_output_paths

FIXTURE_260722_PATH = "tests/fixtures/260722 GTK WIP.xlsx"
FIXTURE_260723_ATX = "tests/fixtures/260723 ATX WIP.xlsx"
FIXTURE_260724_ATX = "tests/fixtures/260724 ATX WIP.xlsx"


def _file_hash(path):
    with open(path, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()


def test_derive_output_paths_uses_date_prefix_and_given_folder():
    today_path = os.path.join("C:", os.sep, "data", "260722 GTK WIP.xlsx")
    output_folder = os.path.join("D:", os.sep, "reports")

    report_path, highlighted_path = derive_output_paths(today_path, output_folder)

    assert report_path == os.path.join("D:", os.sep, "reports", "260722_GTK_WIP_변동리포트.xlsx")
    assert highlighted_path == os.path.join("D:", os.sep, "reports", "260722 GTK WIP_변동표시.xlsx")


def test_derive_output_paths_works_for_atx_filename():
    today_path = os.path.join("C:", os.sep, "data", "260723 ATX WIP.xlsx")
    output_folder = os.path.join("C:", os.sep, "data")

    report_path, highlighted_path = derive_output_paths(today_path, output_folder)

    assert report_path == os.path.join("C:", os.sep, "data", "260723_ATX_WIP_변동리포트.xlsx")
    assert highlighted_path == os.path.join("C:", os.sep, "data", "260723 ATX WIP_변동표시.xlsx")


def test_derive_output_paths_falls_back_when_pattern_does_not_match():
    today_path = os.path.join("C:", os.sep, "data", "random_file.xlsx")
    output_folder = os.path.join("C:", os.sep, "data")

    report_path, highlighted_path = derive_output_paths(today_path, output_folder)

    assert report_path == os.path.join("C:", os.sep, "data", "random_file_WIP_변동리포트.xlsx")
    assert highlighted_path == os.path.join("C:", os.sep, "data", "random_file_변동표시.xlsx")


def test_derive_output_paths_falls_back_when_no_company_token_present():
    # Known accepted trade-off: a date+WIP filename with no company token in
    # between (no real GTK/ATX file looks like this) doesn't match the
    # date+company+WIP regex, so it falls through to the generic fallback,
    # producing a cosmetically odd but harmless "WIP_WIP" filename.
    today_path = os.path.join("C:", os.sep, "data", "260723 WIP.xlsx")
    output_folder = os.path.join("C:", os.sep, "data")

    report_path, highlighted_path = derive_output_paths(today_path, output_folder)

    assert report_path == os.path.join("C:", os.sep, "data", "260723 WIP_WIP_변동리포트.xlsx")
    assert highlighted_path == os.path.join("C:", os.sep, "data", "260723 WIP_변동표시.xlsx")


def test_derive_output_paths_uses_output_folder_not_today_paths_folder():
    today_path = os.path.join("C:", os.sep, "somewhere", "260722 GTK WIP.xlsx")
    output_folder = os.path.join("E:", os.sep, "다른폴더")

    report_path, highlighted_path = derive_output_paths(today_path, output_folder)

    assert report_path == os.path.join("E:", os.sep, "다른폴더", "260722_GTK_WIP_변동리포트.xlsx")
    assert highlighted_path == os.path.join("E:", os.sep, "다른폴더", "260722 GTK WIP_변동표시.xlsx")


STAGE_SUMMARY = {
    "전공정": {"yesterday": 100, "today": 50, "delta": -50},
    "후공정": {"yesterday": 10, "today": 110, "delta": 100},
    "완료": {"yesterday": 0, "today": 0, "delta": 0},
}

COLUMN_LABELS = {9: "Saw", 10: "Die Mount", 12: "Mold"}

LOT_DIFF = {
    "changed_lots": [
        {
            "key": ("m1", "l1", "d1", 0),
            "row_in_today": 7,
            "changes": [{"col": 9, "label": "Saw(I)", "before": 347638, "after": 0}],
        },
        {
            "key": ("m4", "l4", "d4", 0),
            "row_in_today": 10,
            "changes": [
                {"col": 9, "label": "Saw(I)", "before": 100, "after": 0},
                {"col": 10, "label": "Die Mount(J)", "before": 0, "after": 100},
            ],
        },
    ],
    "new_lots": [{"key": ("m2", "l2", "d2", 0), "row_in_today": 8, "values": {9: 123, 12: 456}}],
    "removed_lots": [{"key": ("m3", "l3", "d3", 0), "row_in_yesterday": 9, "values": {9: 789, 12: 321}}],
    "process_columns": [9, 10, 12],
}


def test_build_variance_report_writes_expected_sheets(tmp_path):
    output_path = tmp_path / "report.xlsx"

    build_variance_report(
        STAGE_SUMMARY, LOT_DIFF, str(output_path), COLUMN_LABELS, "#,##0", ("MO", "랏번호", "디바이스")
    )

    wb = openpyxl.load_workbook(str(output_path))
    assert wb.sheetnames == ["요약", "변동랏", "신규랏", "삭제랏"]

    summary_ws = wb["요약"]
    assert summary_ws["A2"].value == "전공정"
    assert summary_ws["B2"].value == 100
    assert summary_ws["C2"].value == 50
    assert summary_ws["D2"].value == -50

    changed_ws = wb["변동랏"]
    assert changed_ws["B2"].value == "l1"
    assert changed_ws["D2"].value == "Saw(I)"
    assert changed_ws["E2"].value == 347638
    assert changed_ws["F2"].value == 0

    # second changed lot has 2 changes -> flattens into rows 3 and 4
    assert changed_ws["B3"].value == "l4"
    assert changed_ws["D3"].value == "Saw(I)"
    assert changed_ws["E3"].value == 100
    assert changed_ws["F3"].value == 0
    assert changed_ws["B4"].value == "l4"
    assert changed_ws["D4"].value == "Die Mount(J)"
    assert changed_ws["E4"].value == 0
    assert changed_ws["F4"].value == 100

    new_ws = wb["신규랏"]
    assert new_ws["D1"].value == "Saw(I)"
    assert new_ws["E1"].value == "Die Mount(J)"
    assert new_ws["B2"].value == "l2"
    assert new_ws["D2"].value == 123
    assert new_ws["E2"].value == 0

    removed_ws = wb["삭제랏"]
    assert removed_ws["D1"].value == "Saw(I)"
    assert removed_ws["B2"].value == "l3"
    assert removed_ws["D2"].value == 789
    assert removed_ws["E2"].value == 0


def test_build_variance_report_applies_readability_formatting(tmp_path):
    output_path = tmp_path / "report.xlsx"

    build_variance_report(
        STAGE_SUMMARY, LOT_DIFF, str(output_path), COLUMN_LABELS, "#,##0", ("MO", "랏번호", "디바이스")
    )

    wb = openpyxl.load_workbook(str(output_path))

    for sheet_name in ["요약", "변동랏", "신규랏", "삭제랏"]:
        ws = wb[sheet_name]
        assert ws.freeze_panes == "A2"
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        assert all(cell.font.bold for cell in header_row if cell.value is not None)
        assert ws.column_dimensions["A"].width is not None
        assert ws.column_dimensions["A"].width > 0

    summary_ws = wb["요약"]
    assert summary_ws["B2"].number_format == "#,##0"
    assert summary_ws["C2"].number_format == "#,##0"
    assert summary_ws["D2"].number_format == "#,##0"

    changed_ws = wb["변동랏"]
    assert changed_ws["E2"].number_format == "#,##0"
    assert changed_ws["F2"].number_format == "#,##0"

    new_ws = wb["신규랏"]
    assert new_ws["D2"].number_format == "#,##0"

    removed_ws = wb["삭제랏"]
    assert removed_ws["D2"].number_format == "#,##0"


def test_build_variance_report_uses_atx_key_labels_not_gtk_hardcoded_headers(tmp_path):
    output_path = tmp_path / "report.xlsx"

    build_variance_report(
        STAGE_SUMMARY,
        LOT_DIFF,
        str(output_path),
        COLUMN_LABELS,
        "#,##0",
        ("웨이퍼랏", "디바이스", "컨트롤랏"),
    )

    wb = openpyxl.load_workbook(str(output_path))

    changed_ws = wb["변동랏"]
    assert changed_ws["A1"].value == "웨이퍼랏"
    assert changed_ws["B1"].value == "디바이스"
    assert changed_ws["C1"].value == "컨트롤랏"

    new_ws = wb["신규랏"]
    assert new_ws["A1"].value == "웨이퍼랏"
    assert new_ws["B1"].value == "디바이스"
    assert new_ws["C1"].value == "컨트롤랏"

    removed_ws = wb["삭제랏"]
    assert removed_ws["A1"].value == "웨이퍼랏"
    assert removed_ws["B1"].value == "디바이스"
    assert removed_ws["C1"].value == "컨트롤랏"


def test_build_variance_report_rejects_key_labels_of_wrong_length(tmp_path):
    output_path = tmp_path / "report.xlsx"

    with pytest.raises(ValueError):
        build_variance_report(
            STAGE_SUMMARY,
            LOT_DIFF,
            str(output_path),
            COLUMN_LABELS,
            "#,##0",
            ("웨이퍼랏", "디바이스"),
        )


def test_build_highlighted_today_file_marks_cells_without_touching_original(tmp_path):
    today_copy = tmp_path / "260722 GTK WIP.xlsx"
    shutil.copyfile(FIXTURE_260722_PATH, today_copy)
    original_hash = _file_hash(today_copy)

    lot_diff = {
        "changed_lots": [
            {
                "key": ("TNT0896622324", "1B3C75", "TMP1202D", 0),
                "row_in_today": 21,
                "changes": [
                    {"col": 9, "label": "Saw(I)", "before": 347638, "after": 0},
                    {"col": 10, "label": "Die Mount(J)", "before": 316800, "after": 664438},
                ],
            },
        ],
        "new_lots": [{"key": ("MNS08M6622326", "1B3C76", "TMP1230", 0), "row_in_today": 22}],
        "removed_lots": [],
    }

    output_path = tmp_path / "260722 GTK WIP_변동표시.xlsx"
    build_highlighted_today_file(str(today_copy), lot_diff, str(output_path), "gtk3387")

    assert _file_hash(today_copy) == original_hash

    wb = openpyxl.load_workbook(str(output_path))
    ws = wb[wb.sheetnames[0]]
    assert ws.cell(row=21, column=9).fill.fgColor.rgb == "FFFF0000"
    assert ws.cell(row=21, column=10).fill.fgColor.rgb == "FFFF0000"
    assert ws.cell(row=21, column=1).fill.fgColor.rgb != "FFFF0000"
    assert ws.cell(row=22, column=1).fill.fgColor.rgb == "FFADD8E6"
    assert ws.cell(row=22, column=38).fill.fgColor.rgb == "FFADD8E6"


def test_build_highlighted_today_file_uses_sheet_name_not_first_sheet(tmp_path):
    wb = openpyxl.Workbook()
    first_ws = wb.active
    first_ws.title = "OtherSheet"
    first_ws["A1"] = "unrelated"
    target_ws = wb.create_sheet("TargetSheet")
    target_ws["A1"] = "data"
    today_path = tmp_path / "multi_sheet.xlsx"
    wb.save(today_path)

    lot_diff = {
        "changed_lots": [
            {"key": ("m1", "l1", "d1", 0), "row_in_today": 1, "changes": [{"col": 1, "label": "A(A)", "before": 1, "after": 2}]},
        ],
        "new_lots": [],
        "removed_lots": [],
    }

    output_path = tmp_path / "multi_sheet_highlighted.xlsx"
    build_highlighted_today_file(str(today_path), lot_diff, str(output_path), "TargetSheet")

    result_wb = openpyxl.load_workbook(str(output_path))
    assert result_wb.sheetnames[0] == "OtherSheet"
    assert result_wb["TargetSheet"].cell(row=1, column=1).fill.fgColor.rgb == "FFFF0000"
    assert result_wb["OtherSheet"].cell(row=1, column=1).fill.fgColor.rgb != "FFFF0000"


def test_atx_end_to_end_report_and_highlight(tmp_path):
    yesterday = parse_atx_wip_sheet(FIXTURE_260723_ATX)
    today = parse_atx_wip_sheet(FIXTURE_260724_ATX)

    stage_summary = compare_stage_summary(yesterday, today)
    lot_diff = compare_lots(yesterday, today)

    assert set(stage_summary.keys()) == {"전공정", "후공정"}
    assert len(lot_diff["changed_lots"]) == 81
    assert len(lot_diff["new_lots"]) == 25
    assert len(lot_diff["removed_lots"]) == 9

    report_path = tmp_path / "report.xlsx"
    build_variance_report(
        stage_summary,
        lot_diff,
        str(report_path),
        today.column_labels,
        today.value_number_format,
        today.key_labels,
    )
    report_wb = openpyxl.load_workbook(str(report_path))
    assert report_wb["요약"]["A2"].value == "전공정"
    assert report_wb["변동랏"]["E2"].number_format == "#,##0.00"
    assert report_wb["변동랏"]["A1"].value == "웨이퍼랏"
    assert report_wb["변동랏"]["B1"].value == "디바이스"
    assert report_wb["변동랏"]["C1"].value == "컨트롤랏"
    assert report_wb["신규랏"]["A1"].value == "웨이퍼랏"
    assert report_wb["신규랏"]["B1"].value == "디바이스"
    assert report_wb["신규랏"]["C1"].value == "컨트롤랏"
    # Rows are written in lot_diff["new_lots"] order, so row 2 corresponds to
    # the first new lot: confirm WAFERLOT/DEVICE/CONTROLLOT values land under
    # their correctly-labeled columns (A=웨이퍼랏, B=디바이스, C=컨트롤랏).
    first_new_key = lot_diff["new_lots"][0]["key"]
    new_ws = report_wb["신규랏"]
    assert (new_ws["A2"].value, new_ws["B2"].value, new_ws["C2"].value) == first_new_key[:3]

    today_copy = tmp_path / "260724 ATX WIP.xlsx"
    shutil.copyfile(FIXTURE_260724_ATX, today_copy)
    highlighted_path = tmp_path / "260724 ATX WIP_변동표시.xlsx"
    build_highlighted_today_file(str(today_copy), lot_diff, str(highlighted_path), today.sheet_name)

    highlighted_wb = openpyxl.load_workbook(str(highlighted_path))
    ws = highlighted_wb[today.sheet_name]
    first_changed = next(
        lot for lot in lot_diff["changed_lots"]
        if lot["key"] == ("BQ8886001A", "GTMP122007", "TPSJ26N009", 0)
    )
    changed_col = first_changed["changes"][0]["col"]
    assert ws.cell(row=first_changed["row_in_today"], column=changed_col).fill.fgColor.rgb == "FFFF0000"

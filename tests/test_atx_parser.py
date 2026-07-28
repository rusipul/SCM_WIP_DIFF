import openpyxl
import pytest

from scm_wip_diff.atx_parser import find_atx_sheet, get_atx_stage_groups, parse_atx_wip_sheet
from scm_wip_diff.parser import ReportFormatError

FIXTURE_260723 = "tests/fixtures/260723 ATX WIP.xlsx"
FIXTURE_260724 = "tests/fixtures/260724 ATX WIP.xlsx"


def test_find_atx_sheet_matches_by_prefix_regardless_of_suffix():
    wb_723 = openpyxl.load_workbook(FIXTURE_260723, data_only=True)
    wb_724 = openpyxl.load_workbook(FIXTURE_260724, data_only=True)

    assert find_atx_sheet(wb_723).title == "KSWIPAY (PKG)"
    assert find_atx_sheet(wb_724).title == "KSWIPAY"


def test_find_atx_sheet_raises_when_no_matching_sheet():
    wb = openpyxl.Workbook()

    with pytest.raises(ReportFormatError):
        find_atx_sheet(wb)


def test_get_atx_stage_groups_uses_header_text_not_merged_cells():
    wb = openpyxl.load_workbook(FIXTURE_260723, data_only=True)
    ws = find_atx_sheet(wb)

    groups = get_atx_stage_groups(ws)

    assert groups["전공정"] == list(range(13, 23))
    assert groups["후공정"] == list(range(24, 47))


def test_get_atx_stage_groups_matches_across_both_fixtures_despite_merge_drift():
    wb_723 = openpyxl.load_workbook(FIXTURE_260723, data_only=True)
    wb_724 = openpyxl.load_workbook(FIXTURE_260724, data_only=True)

    groups_723 = get_atx_stage_groups(find_atx_sheet(wb_723))
    groups_724 = get_atx_stage_groups(find_atx_sheet(wb_724))

    assert groups_723 == groups_724


def test_get_atx_stage_groups_raises_when_required_label_missing():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=2, column=13, value="UNISSUE")
    ws.cell(row=2, column=23, value="FE Total")
    ws.cell(row=2, column=24, value="Ftape")
    # "BE Total" 헤더를 의도적으로 생략해 포맷 변경을 시뮬레이션

    with pytest.raises(ReportFormatError):
        get_atx_stage_groups(ws)


def test_parse_atx_wip_sheet_reads_all_lot_rows():
    parsed = parse_atx_wip_sheet(FIXTURE_260723)

    assert len(parsed.lots) == 125
    assert parsed.sheet_name == "KSWIPAY (PKG)"
    assert parsed.value_number_format == "#,##0.00"


def test_parse_atx_wip_sheet_extracts_process_values_for_first_data_row():
    parsed = parse_atx_wip_sheet(FIXTURE_260723)
    key = ("BQ8886001A", "GTMP122007", "TPSJ26N009", 0)

    assert parsed.rows[key] == 3
    assert parsed.lots[key][13] == 0.0     # UNISSUE
    assert parsed.lots[key][16] == 54.66   # Die_Bond


def test_parse_atx_wip_sheet_disambiguates_duplicate_keys_by_occurrence():
    parsed = parse_atx_wip_sheet(FIXTURE_260723)
    base = ("BQ8873301A", "GTMP17500D", "TPNJ28N009")

    assert parsed.rows[base + (0,)] == 13
    assert parsed.rows[base + (1,)] == 14
    assert parsed.rows[base + (2,)] == 15
    assert parsed.lots[base + (0,)][17] == 22.98   # Wire_Bond
    assert parsed.lots[base + (1,)][17] == 28.61
    assert parsed.lots[base + (2,)][17] == 34.48


def test_parse_atx_wip_sheet_populates_key_labels():
    parsed = parse_atx_wip_sheet(FIXTURE_260723)

    assert parsed.key_labels == ("웨이퍼랏", "디바이스", "컨트롤랏")

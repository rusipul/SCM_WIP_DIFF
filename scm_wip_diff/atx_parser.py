"""Parse ATX WIP Excel reports (KSWIPAY sheet) into structured lot-level data."""

import openpyxl

from scm_wip_diff.parser import ParsedWip, ReportFormatError

ATX_SHEET_PREFIX = "KSWIPAY"
HEADER_ROW = 2
DATA_START_ROW = 3
VALUE_NUMBER_FORMAT = "#,##0.00"


def find_atx_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith(ATX_SHEET_PREFIX):
            return wb[name]
    raise ReportFormatError(f"'{ATX_SHEET_PREFIX}'로 시작하는 시트를 찾을 수 없습니다")


REQUIRED_HEADER_LABELS = ["UNISSUE", "FE Total", "Ftape", "BE Total"]


def _find_label_column(ws, label):
    for col in range(1, ws.max_column + 1):
        if str(ws.cell(row=HEADER_ROW, column=col).value or "").strip() == label:
            return col
    return None


def get_atx_stage_groups(ws):
    positions = {label: _find_label_column(ws, label) for label in REQUIRED_HEADER_LABELS}
    missing = [label for label, col in positions.items() if col is None]
    if missing:
        raise ReportFormatError(f"필수 컬럼 헤더를 찾을 수 없습니다: {missing}")

    unissue, fe_total, ftape, be_total = (positions[label] for label in REQUIRED_HEADER_LABELS)
    return {
        "전공정": list(range(unissue, fe_total)),
        "후공정": list(range(ftape, be_total)),
    }


def parse_atx_wip_sheet(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = find_atx_sheet(wb)

    groups = get_atx_stage_groups(ws)
    process_cols = sorted({c for cols in groups.values() for c in cols})
    labels = {
        col: str(ws.cell(row=HEADER_ROW, column=col).value or "").strip()
        for col in process_cols
    }

    lots = {}
    rows = {}
    occurrence = {}
    r = DATA_START_ROW
    while ws.cell(row=r, column=1).value is not None:
        device = str(ws.cell(row=r, column=3).value or "").strip()
        wafer_lot = str(ws.cell(row=r, column=5).value or "").strip()
        control_lot = str(ws.cell(row=r, column=6).value or "").strip()
        base_key = (wafer_lot, device, control_lot)
        idx = occurrence.get(base_key, 0)
        occurrence[base_key] = idx + 1
        key = (wafer_lot, device, control_lot, idx)

        values = {}
        for col in process_cols:
            v = ws.cell(row=r, column=col).value
            values[col] = float(v) if isinstance(v, (int, float)) else 0.0

        lots[key] = values
        rows[key] = r
        r += 1

    return ParsedWip(
        column_labels=labels,
        stage_groups=groups,
        lots=lots,
        rows=rows,
        sheet_name=ws.title,
        value_number_format=VALUE_NUMBER_FORMAT,
        key_labels=("웨이퍼랏", "디바이스", "컨트롤랏"),
    )

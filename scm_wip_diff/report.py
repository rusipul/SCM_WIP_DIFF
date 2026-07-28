"""Generate the variance report workbook and the highlighted today-file copy."""
import os
import re
import shutil

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

RED_FILL = PatternFill(fill_type="solid", fgColor="FFFF0000")
NEW_LOT_FILL = PatternFill(fill_type="solid", fgColor="FFADD8E6")

STAGE_ORDER = ["전공정", "후공정", "완료"]


def derive_output_paths(today_path):
    folder = os.path.dirname(today_path)
    basename = os.path.basename(today_path)
    name, ext = os.path.splitext(basename)
    match = re.match(r"^(\d{6})\s+(\S+)\s+WIP", basename)
    if match:
        date_prefix, company = match.group(1), match.group(2)
        report_name = f"{date_prefix}_{company}_WIP_변동리포트.xlsx"
    else:
        report_name = f"{name}_WIP_변동리포트.xlsx"
    report_path = os.path.join(folder, report_name)
    highlighted_path = os.path.join(folder, f"{name}_변동표시{ext}")
    return report_path, highlighted_path


HEADER_FONT = Font(bold=True)
MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 40


def _process_column_headers(column_labels, process_columns):
    return [f"{column_labels.get(col, '')}({get_column_letter(col)})" for col in process_columns]


def _apply_readability_formatting(ws, number_format_columns, number_format):
    for cell in ws[1]:
        if cell.value is not None:
            cell.font = HEADER_FONT
    ws.freeze_panes = "A2"

    for col_cells in ws.columns:
        col_letter = col_cells[0].column_letter
        max_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
        ws.column_dimensions[col_letter].width = min(
            max(max_len + 2, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH
        )

    for col in number_format_columns:
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = number_format


def build_variance_report(
    stage_summary, lot_diff, output_path, column_labels, value_number_format, key_labels
):
    if len(key_labels) != 3:
        raise ValueError(f"key_labels must have exactly 3 elements, got {len(key_labels)}: {key_labels}")

    wb = openpyxl.Workbook()

    summary_ws = wb.active
    summary_ws.title = "요약"
    summary_ws.append(["단계", "어제", "오늘", "증감"])
    for stage in STAGE_ORDER:
        if stage in stage_summary:
            s = stage_summary[stage]
            summary_ws.append([stage, s["yesterday"], s["today"], s["delta"]])
    summary_ws.append([])
    summary_ws.append(["변경된 랏 수", len(lot_diff["changed_lots"])])
    summary_ws.append(["신규 랏 수", len(lot_diff["new_lots"])])
    summary_ws.append(["삭제된 랏 수", len(lot_diff["removed_lots"])])
    _apply_readability_formatting(summary_ws, number_format_columns=[2, 3, 4], number_format=value_number_format)

    changed_ws = wb.create_sheet("변동랏")
    changed_ws.append(list(key_labels) + ["변경컬럼", "어제값", "오늘값"])
    for lot in lot_diff["changed_lots"]:
        key1, key2, key3, _ = lot["key"]
        for change in lot["changes"]:
            changed_ws.append([key1, key2, key3, change["label"], change["before"], change["after"]])
    _apply_readability_formatting(changed_ws, number_format_columns=[5, 6], number_format=value_number_format)

    process_columns = lot_diff["process_columns"]
    process_headers = _process_column_headers(column_labels, process_columns)
    process_number_format_columns = list(range(4, 4 + len(process_headers)))

    new_ws = wb.create_sheet("신규랏")
    new_ws.append(list(key_labels) + process_headers)
    for lot in lot_diff["new_lots"]:
        key1, key2, key3, _ = lot["key"]
        values = lot["values"]
        new_ws.append([key1, key2, key3] + [values.get(col, 0) for col in process_columns])
    _apply_readability_formatting(new_ws, number_format_columns=process_number_format_columns, number_format=value_number_format)

    removed_ws = wb.create_sheet("삭제랏")
    removed_ws.append(list(key_labels) + process_headers)
    for lot in lot_diff["removed_lots"]:
        key1, key2, key3, _ = lot["key"]
        values = lot["values"]
        removed_ws.append([key1, key2, key3] + [values.get(col, 0) for col in process_columns])
    _apply_readability_formatting(removed_ws, number_format_columns=process_number_format_columns, number_format=value_number_format)

    wb.save(output_path)


def build_highlighted_today_file(today_path, lot_diff, output_path, sheet_name):
    shutil.copyfile(today_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]

    for lot in lot_diff["changed_lots"]:
        row = lot["row_in_today"]
        for change in lot["changes"]:
            ws.cell(row=row, column=change["col"]).fill = RED_FILL

    for lot in lot_diff["new_lots"]:
        row = lot["row_in_today"]
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).fill = NEW_LOT_FILL

    wb.save(output_path)
